from openpyxl import load_workbook,Workbook
import matplotlib.pyplot as plt
import numpy as np

wb = load_workbook('NLP-Mallet_Output_Composition_7.xlsx')

sheet = wb['NLP-Mallet_Output_Composition_7']
dic_ = {}
for row in sheet:
	total = 0.0
	dic_[int(row[1].value)] = {}
	for i in range(2,len(row),2):
		dic_[int(row[1].value)][int(row[i].value)]=row[i+1].value
data = []

for each_chapter in dic_:
	list_ = []
	for each_story in dic_[each_chapter]:
		list_.append(dic_[each_chapter][each_story])
	data.append(list_)
plt_data = np.array(data).transpose()

wb_1 = Workbook()
file_name = 'Story_Line_xlsx'
ws = wb_1.active

for row in range(1,len(plt_data)+1):
	for cell in range(1,len(plt_data[row-1])+1):
		_ = ws.cell(column=row, row=cell, value=plt_data[row-1][cell-1])
wb_1.save(filename = file_name)
"""
x = range(1,75)

plt.plot(x,plt_data[0],'b--',label = 'story_0')
plt.plot(x,plt_data[1],'g--',label = 'story_1')
plt.plot(x,plt_data[2],'r--',label = 'story_2')
plt.plot(x,plt_data[3],'c--',label = 'story_3')
plt.plot(x,plt_data[4],'m--',label = 'story_4')
plt.plot(x,plt_data[5],'y--',label = 'story_5')
plt.plot(x,plt_data[6],'k--',label = 'story_6')
plt.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
#plt.figure(figsize=(8,6))
#plt.savefig('Seven_Story_Lines.png')
plt.show()


for i in range(0,7):
	plt.figure(figsize=(8,6))
	counter = 0
	for j in range(0,7):
		if i!=j:
			counter+=1
			plt.subplot(320+counter)
			plt.plot(x,plt_data[i],'r--',x,plt_data[j],'b--')
			title = 'Story %d x Story %s'%(i,j)
			plt.title(title)
			plt.subplots_adjust(hspace = 0.5)
	#plt.show()
	title = 'Story Line %d.png'%(i)
	plt.savefig(title)
"""
